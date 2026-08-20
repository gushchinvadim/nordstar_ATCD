# reporting/services/rauc_export.py
import os
import re
import hashlib
from datetime import timedelta
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom

from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from execution.models import Enrollment
from reporting.models import RegulatoryReport, RegulatoryReportItem


class RAUCExportService:
    """Сервис генерации отчёта для РАУЦ (Excel и XML)"""

    HEADERS = [
        'SURNAME', 'NAME', 'PATRONYMIC', 'DBIRTH',
        'PROG_ID', 'MOD_ID',
        'DBEGINEXT', 'DBEGIN', 'DEND',
        'NGROUP', 'ADDR_ID', 'DCAT_ID',
        'NDOC', 'DDOC',
        'FPTITLE_ID', 'TPTITLE_ID',
        'DENDDOC', 'DEPT_ID'
    ]

    def __init__(self, group, user=None):
        self.group = group
        self.user = user  # Пользователь, который генерирует отчёт

        # Инициализация пути к папке группы
        year = str(group.start_date.year) if group.start_date else 'unknown_year'
        module_code = re.sub(r'[^\w\-]', '_', group.module.code) if group.module else 'unknown_module'

        self.folder_path = os.path.join(
            settings.MEDIA_ROOT,
            'documents', str(year), 'groups', module_code, group.assigned_number, 'reports'
        )
        os.makedirs(self.folder_path, exist_ok=True)

    def get_data(self):
        """Собирает и валидирует данные"""
        enrollments = Enrollment.objects.filter(
            group=self.group,
            status='completed'
        ).select_related(
            'student', 'group', 'group__module', 'group__module__course',
            'group__location', 'group__curator', 'group__director'
        ).prefetch_related('certificates')

        rows = []
        errors = []

        for enrollment in enrollments:
            row, row_errors = self._map_row(enrollment)
            if row:
                rows.append(row)
            errors.extend(row_errors)

        return rows, errors

    def _map_row(self, enrollment):
        """Точный маппинг по образцу РАУЦ.xlsx"""
        errors = []
        s = enrollment.student
        g = enrollment.group
        module = g.module
        course = module.course if module else None
        loc = g.location

        cert = enrollment.certificates.first()
        if not cert:
            errors.append(f"Студент {s.surname} {s.name}: нет сертификата")
            return None, errors

        # FPTITLE_ID и TPTITLE_ID из rauts_id куратора и директора
        fptitle_id = getattr(g.curator, 'rauts_id', '') if g.curator else ''
        tptitle_id = getattr(g.director, 'rauts_id', '') if g.director else ''

        if not fptitle_id:
            errors.append(f"У куратора '{g.curator}' не заполнен rauts_id")
        if not tptitle_id:
            errors.append(f"У директора '{g.director}' не заполнен rauts_id")

        # DENDDOC: расчёт даты окончания действия
        denndoc = ''
        if cert.issue_date:
            if module and getattr(module, 'validity_period', None):
                expiry_date = cert.issue_date + relativedelta(months=module.validity_period)
                denndoc = expiry_date.strftime('%d.%m.%Y')
            else:
                denndoc = (cert.issue_date + timedelta(days=365)).strftime('%d.%m.%Y')
                errors.append(f"Модуль '{module.code}': не указан validity_period, использован fallback +1 год")

        # Валидация обязательных полей
        if not s.dob: errors.append(f"Студент {s.surname}: нет даты рождения")
        if not course or not course.prog_id: errors.append(f"Программа: не заполнен prog_id")
        if not module or not module.mod_id: errors.append(f"Модуль: не заполнен mod_id")
        if not loc or not loc.addr: errors.append(f"Локация: не заполнен addr (код РАУЦ)")
        if not getattr(s, 'dcat_id', None): errors.append(f"Студент {s.surname}: не заполнен dcat_id")

        # Формирование строки строго по шаблону
        row = {
            'SURNAME': s.surname,
            'NAME': s.name,
            'PATRONYMIC': s.patronymic or '',
            'DBIRTH': s.dob.strftime('%d.%m.%Y') if s.dob else '',
            'PROG_ID': course.prog_id if course else '',
            'MOD_ID': module.mod_id if module else '',
            'DBEGINEXT': g.start_date.strftime('%d.%m.%Y') if g.start_date else '',
            'DBEGIN': g.start_face_to_face.strftime('%d.%m.%Y') if g.start_face_to_face else '',
            'DEND': g.end_date.strftime('%d.%m.%Y') if g.end_date else '',
            'NGROUP': f"{g.assigned_number} - {g.application}" if g.application else g.assigned_number,
            'ADDR_ID': loc.addr if loc else '',
            'DCAT_ID': getattr(s, 'dcat_id', '') or '',
            'NDOC': cert.number,
            'DDOC': cert.issue_date.strftime('%d.%m.%Y') if cert.issue_date else '',
            'FPTITLE_ID': fptitle_id,
            'TPTITLE_ID': tptitle_id,
            'DENDDOC': denndoc,
            'DEPT_ID': loc.dept if loc else '',
        }

        return row, errors

    def generate_excel(self):
        """Генерирует Excel-файл и сохраняет в папку группы"""
        rows, errors = self.get_data()
        if not rows:
            raise ValueError(f"Нет завершённых студентов с сертификатами в группе {self.group.assigned_number}")

        wb = Workbook()
        ws = wb.active
        ws.title = "РАУЦ"

        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for col_idx, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(self.HEADERS, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ''))
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        for col_idx in range(1, len(self.HEADERS) + 1):
            max_length = max(len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, len(rows) + 2))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

        filename = f"РАУЦ_{self.group.assigned_number}.xlsx"
        filepath = os.path.join(self.folder_path, filename)
        wb.save(filepath)

        rel_path = os.path.relpath(filepath, settings.MEDIA_ROOT).replace('\\', '/')
        return filepath, rel_path, errors

    def generate_xml(self):
        """Генерирует XML-файл и сохраняет в папку группы"""
        rows, errors = self.get_data()
        if not rows:
            raise ValueError(f"Нет завершённых студентов с сертификатами в группе {self.group.assigned_number}")

        root = Element('RAUCReport')
        root.set('group', self.group.assigned_number)

        for row in rows:
            record = SubElement(root, 'Record')
            for header in self.HEADERS:
                field = SubElement(record, header)
                field.text = str(row.get(header, ''))

        xml_string = tostring(root, encoding='unicode')
        pretty_xml = minidom.parseString(xml_string).toprettyxml(indent='  ', encoding='UTF-8')

        filename = f"РАУЦ_{self.group.assigned_number}.xml"
        filepath = os.path.join(self.folder_path, filename)

        with open(filepath, 'wb') as f:
            f.write(pretty_xml)

        rel_path = os.path.relpath(filepath, settings.MEDIA_ROOT).replace('\\', '/')
        return filepath, rel_path, errors

    def save_to_database(self, excel_path=None, xml_path=None, file_format='excel'):
        from reporting.models import RegulatoryReport, RegulatoryReportItem
        from people.models import Student
        import hashlib

        rows, errors = self.get_data()
        status = 'draft' if errors else 'generated'
        format_label = 'Excel' if file_format == 'excel' else 'XML'
        title = f"РАУЦ ({format_label}) — группа {self.group.assigned_number}"

        report = RegulatoryReport.objects.create(
            report_type='rauc', title=title, status=status, created_by=self.user
        )
        report.groups.add(self.group)

        if excel_path and os.path.exists(excel_path):
            with open(excel_path, 'rb') as f:
                report.excel_file.save(f"РАУЦ_{self.group.assigned_number}.xlsx", f, save=False)

        if xml_path and os.path.exists(xml_path):
            with open(xml_path, 'rb') as f:
                xml_content = f.read()
                report.xml_file.save(f"РАУЦ_{self.group.assigned_number}.xml", f, save=False)
                report.xml_file_hash = hashlib.sha256(xml_content).hexdigest()

        report.save()

        for row in rows:
            surname = row.get('SURNAME', '')
            name = row.get('NAME', '')
            patronymic = row.get('PATRONYMIC', '')

            try:
                student_obj = Student.objects.get(
                    surname=surname,
                    name=name,
                    patronymic=patronymic
                )

                enrollment_obj = Enrollment.objects.filter(
                    group=self.group,
                    student=student_obj
                ).first()

                certificate_obj = enrollment_obj.certificates.first() if enrollment_obj else None

                row_errors = []
                if not row.get('PROG_ID'): row_errors.append("Не заполнен PROG_ID")
                if not row.get('MOD_ID'): row_errors.append("Не заполнен MOD_ID")
                if not row.get('DBIRTH'): row_errors.append("Не заполнена дата рождения")

                RegulatoryReportItem.objects.create(
                    report=report,
                    enrollment=enrollment_obj,
                    student=student_obj,  # <-- ЭТО ИСПРАВЛЯЕТ ОШИБКУ NOT NULL
                    certificate=certificate_obj,
                    is_valid=len(row_errors) == 0,
                    validation_error='; '.join(row_errors),
                    payload=row
                )

            except Student.DoesNotExist:
                errors.append(f"Критическая ошибка: Студент {surname} {name} {patronymic} не найден в базе данных")

        if errors and report.status == 'generated':
            report.status = 'draft'
            report.save()

        return report