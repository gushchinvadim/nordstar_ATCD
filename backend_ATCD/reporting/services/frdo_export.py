# reporting/services/frdo_export.py
import os
import re
import math
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from execution.models import Enrollment


class FRDOExportService:
    """Сервис генерации отчёта для ФИС ФРДО с поддержкой нескольких шаблонов"""

    HEADERS_ATTENDANT = [
        'Вид документа', 'Статус документа', 'Подтверждение утраты', 'Подтверждение обмена',
        'Подтверждение уничтожения', 'Серия документа', 'Номер документа', 'Дата выдачи документа',
        'Регистрационный номер', 'Программа профессионального обучения, направление подготовки',
        'Наименование программы профессионального обучения', 'Наименование профессий рабочих, должностей служащих',
        'Присвоенный квалификационный разряд, класс, категория (при наличии)', 'Год начала обучения',
        'Год окончания обучения', 'Срок обучения, часов', 'Фамилия получателя', 'Имя получателя',
        'Отчество получателя', 'Дата рождения получателя', 'Пол получателя', 'СНИЛС',
        'Гражданство получателя (код страны по ОКСМ)', 'Форма обучения', 'Источник финансирования обучения',
        'Форма получения образования на момент прекращения образовательных отношений',
        'Наименование документа об образовании (оригинала)', 'Серия (оригинала)', 'Номер (оригинала)',
        'Регистрационный N (оригинала)', 'Дата выдачи (оригинала)', 'Фамилия получателя (оригинала)',
        'Имя получателя (оригинала)', 'Отчество получателя (оригинала)', 'Номер документа для изменения',
    ]

    HEADERS_PILOT = [
        'Вид документа', 'Статус документа', 'Подтверждение утраты', 'Подтверждение обмена',
        'Подтверждение уничтожения', 'Серия документа', 'Номер документа', 'Дата выдачи документа',
        'Регистрационный номер',
        'Дополнительная профессиональная программа (повышение квалификации/ профессиональная переподготовка)',
        'Наименование дополнительной профессиональной программы', 'Наименование области профессиональной деятельности',
        'Укрупненные группы специальностей', 'Наименование квалификации, профессии, специальности',
        'Уровень образования ВО/СПО', 'Фамилия указанная в дипломе о ВО или СПО', 'Серия документа о ВО/СПО',
        'Номер документа о ВО/СПО', 'Год начала обучения (для документа о квалификации)',
        'Год окончания обучения (для документа о квалификации)', 'Срок обучения, часов (для документа о квалификации)',
        'Фамилия получателя', 'Имя получателя', 'Отчество получателя', 'Дата рождения получателя',
        'Пол получателя', 'СНИЛС', 'Форма обучения', 'Источник финансирования обучения',
        'Форма получения образования на момент прекращения образовательных отношений',
        'Гражданство получателя (код страны по ОКСМ)', 'Наименование документа об образовании (оригинала)',
        'Серия (оригинала)', 'Номер (оригинала)', 'Регистрационный N (оригинала)', 'Дата выдачи (оригинала)',
        'Фамилия получателя (оригинала)', 'Имя получателя (оригинала)', 'Отчество получателя (оригинала)',
        'Номер документа для изменения',
    ]

    FRDO_DOCUMENT_TYPES = {
        'credential': 'Удостоверение о повышении квалификации',
        'witness': 'Свидетельство о профессии рабочего, должности служащего',
        'diploma': 'Диплом о профессиональной переподготовке',
        'reference': 'Справка об обучении или о периоде обучения',
        'certificate': 'Сертификат',
    }

    def __init__(self, group, user=None):
        self.group = group
        self.user = user
        self.course = group.module.course if group.module else None
        self.template_type = self.course.frdo_template_type if self.course else 'flight_attendant'

        year = str(group.start_date.year) if group.start_date else 'unknown_year'
        module_code = re.sub(r'[^\w\-]', '_', group.module.code) if group.module else 'unknown_module'

        # Заменяем слэш и другие запрещённые символы в номере группы
        safe_group_number = re.sub(r'[^\w\.\-]', '_', group.assigned_number)

        self.folder_path = os.path.join(
            settings.MEDIA_ROOT, 'documents', str(year), 'groups', module_code, safe_group_number, 'reports'
        )
        os.makedirs(self.folder_path, exist_ok=True)

    def get_headers(self):
        return self.HEADERS_PILOT if self.template_type == 'pilot_engineer' else self.HEADERS_ATTENDANT

    def get_data(self):
        enrollments = Enrollment.objects.filter(
            group=self.group, status='completed'
        ).select_related(
            'student', 'group', 'group__module', 'group__module__course', 'group__location'
        ).prefetch_related('certificates')

        rows, errors = [], []
        for enrollment in enrollments:
            row, row_errors = self._map_row(enrollment)
            if row:
                rows.append(row)
            errors.extend(row_errors)
        return rows, errors

    def _format_snils(self, student, errors):
        raw_snils = str(getattr(student, 'snils', '') or '').strip()
        digits = ''.join(filter(str.isdigit, raw_snils))
        if len(digits) == 11:
            return f"{digits[:3]}-{digits[3:6]}-{digits[6:9]} {digits[9:]}"
        else:
            errors.append(f"Студент {student.surname}: некорректный СНИЛС (ожидается 11 цифр, сейчас {len(digits)})")
            return raw_snils

    def _get_common_data(self, enrollment, errors):
        s = enrollment.student
        g = enrollment.group
        module = g.module
        course = self.course
        cert = enrollment.certificates.first()

        if not cert:
            errors.append(f"Студент {s.surname} {s.name}: нет сертификата")
            return None

        # Серия и номер
        # serial_number — это серия документа для ФРДО (по старому: assigned_number)
        # Если не заполнено, используем assigned_number как fallback
        doc_series = g.serial_number or g.assigned_number or ''
        doc_num = f"{g.application}-{enrollment.number_in_group}" if g.application else str(enrollment.number_in_group)
        doc_number = f"{doc_series}-{doc_num}" if doc_series else doc_num

        # Годы обучения
        year_start = g.start_date.year if g.start_date else ''

        if enrollment.completed_at:
            year_end = enrollment.completed_at.year
        elif cert.issue_date:
            year_end = cert.issue_date.year
            errors.append(f"Студент {s.surname}: не заполнена completed_at")
        else:
            year_end = ''
            errors.append(f"Студент {s.surname}: не заполнены completed_at и issue_date")

        # === Дата выдачи документа (дата приказа) ===
        issue_date = enrollment.order_out_date
        if not issue_date:
            errors.append(f"Студент {s.surname}: не заполнена order_out_date")
            issue_date = cert.issue_date  # fallback
        # ==========================================

        duration_hours = math.ceil(float(module.duration)) if module and module.duration else 0

        # Пол
        sex_value = getattr(s, 'sex', None) or ''
        sex_map = {'male': 'Мужской', 'female': 'Женский', 'М': 'Мужской', 'Ж': 'Женский', 'm': 'Мужской',
                   'f': 'Женский', 'M': 'Муж', 'F': 'Жен'}
        sex = sex_map.get(str(sex_value).lower().strip(), '')
        if not sex:
            errors.append(f"Студент {s.surname}: не заполнен или неверно указан пол")

        formatted_snils = self._format_snils(s, errors)

        form_of_education = course.form_of_education if course and course.form_of_education else ''
        if not form_of_education:
            errors.append(f"Программа: не заполнена форма обучения")

        if not s.dob: errors.append(f"Студент {s.surname}: не заполнена дата рождения")
        if not g.serial_number: errors.append(f"Группа: не заполнена серия документа (serial_number)")

        return {
            's': s, 'g': g, 'module': module, 'course': course, 'cert': cert,
            'doc_series': doc_series, 'doc_num': doc_num, 'doc_number': doc_number,
            'year_start': year_start, 'year_end': year_end, 'duration_hours': duration_hours,
            'sex': sex, 'formatted_snils': formatted_snils, 'form_of_education': form_of_education,
            'issue_date': issue_date,  # ← ДОБАВЛЕНО
            'frdo_doc_type': self.FRDO_DOCUMENT_TYPES.get(cert.certificate_type,
                                                          'Удостоверение о повышении квалификации')
        }

    def _map_row(self, enrollment):
        errors = []
        data = self._get_common_data(enrollment, errors)
        if not data:
            return None, errors

        if self.template_type == 'pilot_engineer':
            return self._map_row_pilot(data, errors)
        else:
            return self._map_row_attendant(data, errors)

    def _map_row_attendant(self, data, errors):
        s, g, module, course, cert = data['s'], data['g'], data['module'], data['course'], data['cert']
        row = {
            'Вид документа': data['frdo_doc_type'],
            'Статус документа': 'Оригинал',
            'Подтверждение утраты': 'Нет', 'Подтверждение обмена': 'Нет', 'Подтверждение уничтожения': 'Нет',
            'Серия документа': data['doc_series'],
            'Номер документа': data['doc_num'],
            'Дата выдачи документа': data['issue_date'].strftime('%d.%m.%Y') if data['issue_date'] else '',
            'Регистрационный номер': data['doc_number'],
            'Программа профессионального обучения, направление подготовки': 'Программа повышения квалификации рабочих, служащих',
            'Наименование программы профессионального обучения': f"{course.title} — {module.title}" if course and module else (
                module.title if module else ''),
            'Наименование профессий рабочих, должностей служащих': 'Бортпроводник',
            'Присвоенный квалификационный разряд, класс, категория (при наличии)': 'Нет',
            'Год начала обучения': str(data['year_start']) if data['year_start'] else '',
            'Год окончания обучения': str(data['year_end']) if data['year_end'] else '',  # ← ИСПРАВЛЕНО
            'Срок обучения, часов': str(data['duration_hours']) if data['duration_hours'] else '0',
            'Фамилия получателя': s.surname,
            'Имя получателя': s.name,
            'Отчество получателя': s.patronymic or '',
            'Дата рождения получателя': s.dob.strftime('%d.%m.%Y') if s.dob else '',
            'Пол получателя': data['sex'],
            'СНИЛС': data['formatted_snils'],
            'Гражданство получателя (код страны по ОКСМ)': str(getattr(s, 'citizenship_code', '') or '643'),
            'Форма обучения': data['form_of_education'],
            'Источник финансирования обучения': 'Платное обучение',
            'Форма получения образования на момент прекращения образовательных отношений': 'в образовательной организации',
            'Наименование документа об образовании (оригинала)': '', 'Серия (оригинала)': '', 'Номер (оригинала)': '',
            'Регистрационный N (оригинала)': '', 'Дата выдачи (оригинала)': '', 'Фамилия получателя (оригинала)': '',
            'Имя получателя (оригинала)': '', 'Отчество получателя (оригинала)': '',
            'Номер документа для изменения': '',
        }
        return row, errors

    def _map_row_pilot(self, data, errors):
        s, g, module, course, cert = data['s'], data['g'], data['module'], data['course'], data['cert']
        row = {
            'Вид документа': data['frdo_doc_type'],
            'Статус документа': 'Оригинал',
            'Подтверждение утраты': 'Нет', 'Подтверждение обмена': 'Нет', 'Подтверждение уничтожения': 'Нет',
            'Серия документа': data['doc_series'],
            'Номер документа': data['doc_num'],
            'Дата выдачи документа': data['issue_date'].strftime('%d.%m.%Y') if data['issue_date'] else '',
            'Регистрационный номер': data['doc_number'],
            'Дополнительная профессиональная программа (повышение квалификации/ профессиональная переподготовка)': 'Повышение квалификации',
            'Наименование дополнительной профессиональной программы': f"{course.title} — {module.title}" if course and module else (
                module.title if module else ''),
            'Наименование области профессиональной деятельности': 'Транспорт',
            'Укрупненные группы специальностей': '',
            'Наименование квалификации, профессии, специальности': str(s.profession) if s.profession else 'нет',
            'Уровень образования ВО/СПО': '',
            'Фамилия указанная в дипломе о ВО или СПО': '',
            'Серия документа о ВО/СПО': '', 'Номер документа о ВО/СПО': '',
            'Год начала обучения (для документа о квалификации)': str(data['year_start']) if data['year_start'] else '',
            'Год окончания обучения (для документа о квалификации)': str(data['year_end']) if data['year_end'] else '',
            # ← ИСПРАВЛЕНО
            'Срок обучения, часов (для документа о квалификации)': str(data['duration_hours']) if data[
                'duration_hours'] else '0',
            'Фамилия получателя': s.surname,
            'Имя получателя': s.name,
            'Отчество получателя': s.patronymic or '',
            'Дата рождения получателя': s.dob.strftime('%d.%m.%Y') if s.dob else '',
            'Пол получателя': data['sex'],
            'СНИЛС': data['formatted_snils'],
            'Форма обучения': data['form_of_education'],
            'Источник финансирования обучения': 'Платное обучение',
            'Форма получения образования на момент прекращения образовательных отношений': 'в образовательной организации',
            'Гражданство получателя (код страны по ОКСМ)': str(getattr(s, 'citizenship_code', '') or '643'),
            'Наименование документа об образовании (оригинала)': '', 'Серия (оригинала)': '', 'Номер (оригинала)': '',
            'Регистрационный N (оригинала)': '', 'Дата выдачи (оригинала)': '', 'Фамилия получателя (оригинала)': '',
            'Имя получателя (оригинала)': '', 'Отчество получателя (оригинала)': '',
            'Номер документа для изменения': '',
        }
        return row, errors

    def generate_excel(self):
        rows, errors = self.get_data()
        if not rows:
            raise ValueError(f"Нет завершённых студентов с сертификатами в группе {self.group.assigned_number}")

        wb = Workbook()
        ws = wb.active
        ws.title = "ФРДО"

        headers = self.get_headers()
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ''))
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        for col_idx in range(1, len(headers) + 1):
            max_length = max(len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, len(rows) + 2))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

        safe_group_number = re.sub(r'[^\w\.\-]', '_', self.group.assigned_number)
        filename = f"ФРДО_{safe_group_number}.xlsx"
        filepath = os.path.join(self.folder_path, filename)
        wb.save(filepath)

        rel_path = os.path.relpath(filepath, settings.MEDIA_ROOT).replace('\\', '/')
        return filepath, rel_path, errors

    def save_to_database(self, excel_path=None, file_format='excel'):
        from reporting.models import RegulatoryReport, RegulatoryReportItem
        from people.models import Student
        import hashlib

        rows, errors = self.get_data()
        status = 'draft' if errors else 'generated'
        title = f"ФРДО ({file_format.upper()}) — группа {self.group.assigned_number}"

        report = RegulatoryReport.objects.create(
            report_type='frdo', title=title, status=status, created_by=self.user
        )
        report.groups.add(self.group)

        safe_group_number = re.sub(r'[^\w\.\-]', '_', self.group.assigned_number)

        if excel_path and os.path.exists(excel_path):
            with open(excel_path, 'rb') as f:
                report.excel_file.save(f"ФРДО_{safe_group_number}.xlsx", f, save=False)
        report.save()

        for row in rows:
            surname = row.get('Фамилия получателя', '')
            name = row.get('Имя получателя', '')
            patronymic = row.get('Отчество получателя', '')

            try:
                student_obj = Student.objects.get(
                    surname=surname,
                    name=name,
                    patronymic=patronymic
                )

                enrollment_obj = Enrollment.objects.filter(
                    group=self.group,
                    student=student_obj
                ).first()

                certificate_obj = enrollment_obj.certificates.first() if enrollment_obj else None

                row_errors = []
                snils_digits = ''.join(filter(str.isdigit, row.get('СНИЛС', '')))
                if len(snils_digits) != 11:
                    row_errors.append("Некорректный СНИЛС")
                if not row.get('Дата рождения получателя'):
                    row_errors.append("Не заполнена дата рождения")
                if not row.get('Год окончания обучения'):
                    row_errors.append("Не заполнен год окончания обучения")

                RegulatoryReportItem.objects.create(
                    report=report,
                    enrollment=enrollment_obj,
                    student=student_obj,
                    certificate=certificate_obj,
                    is_valid=len(row_errors) == 0,
                    validation_error='; '.join(row_errors),
                    payload=row
                )

            except Student.DoesNotExist:
                errors.append(f"Критическая ошибка: Студент {surname} {name} {patronymic} не найден в базе данных")

        if errors and report.status == 'generated':
            report.status = 'draft'
            report.save()

        return report